import os
import re
import io
import tempfile
import cv2
import numpy as np
import pandas as pd
import pymupdf
import camelot
import streamlit as st
from rapidocr import RapidOCR
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from classify_pipeline import (
    classify_excel_dataframe,
    classify_and_save_artifacts,
    get_classification_engine,
    clean_description,
    SUB_TAXONOMY,
    extract_transaction_mode,
    extract_transaction_direction,
    extract_transaction_purpose,
    extract_bank_institution,
    extract_counterparty
)

# ─────────────────────────────────────────────────────────────────────────────
# Page Config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Transaction Classifier",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# Global CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    /* Dark gradient background */
    .stApp {
        background: linear-gradient(135deg, #0a0e1a 0%, #0d1b2a 40%, #0f2744 100%);
        color: #e8edf5;
    }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0d1b2a 0%, #091525 100%);
        border-right: 1px solid rgba(99,179,237,0.15);
    }
    section[data-testid="stSidebar"] * {
        color: #c8d8ea !important;
    }

    /* Hero banner */
    .hero-banner {
        background: linear-gradient(135deg, #1a3a5c 0%, #0e2744 50%, #152d4a 100%);
        border: 1px solid rgba(99,179,237,0.25);
        border-radius: 20px;
        padding: 2.5rem 2rem;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 8px 32px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.05);
    }
    .hero-banner h1 {
        font-size: 2.6rem;
        font-weight: 800;
        background: linear-gradient(90deg, #63b3ed, #90cdf4, #4299e1);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin: 0 0 0.5rem 0;
    }
    .hero-banner p {
        font-size: 1.05rem;
        color: #8fb8d8;
        margin: 0;
        font-weight: 400;
    }

    /* Cards */
    .metric-card {
        background: linear-gradient(135deg, #162030 0%, #1a2d44 100%);
        border: 1px solid rgba(99,179,237,0.2);
        border-radius: 14px;
        padding: 1.4rem 1.2rem;
        text-align: center;
        box-shadow: 0 4px 20px rgba(0,0,0,0.3);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .metric-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 28px rgba(99,179,237,0.15);
    }
    .metric-card .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #63b3ed;
        line-height: 1;
    }
    .metric-card .metric-label {
        font-size: 0.78rem;
        font-weight: 500;
        color: #7a9eba;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin-top: 0.4rem;
    }

    /* Step badge */
    .step-badge {
        background: linear-gradient(135deg, #2b6cb0, #2c5282);
        color: #bee3f8;
        border-radius: 50px;
        padding: 0.25rem 0.9rem;
        font-size: 0.78rem;
        font-weight: 600;
        letter-spacing: 0.05em;
        display: inline-block;
        margin-bottom: 0.5rem;
    }

    /* Section headers */
    .section-header {
        font-size: 1.1rem;
        font-weight: 700;
        color: #90cdf4;
        border-left: 3px solid #2b6cb0;
        padding-left: 0.8rem;
        margin: 1.5rem 0 1rem 0;
        letter-spacing: 0.02em;
    }

    /* Buttons */
    .stButton > button {
        background: linear-gradient(135deg, #2b6cb0 0%, #1a4a7a 100%);
        color: white;
        border: 1px solid rgba(99,179,237,0.4);
        border-radius: 10px;
        font-weight: 600;
        font-size: 0.95rem;
        padding: 0.6rem 1.5rem;
        transition: all 0.2s ease;
        box-shadow: 0 4px 12px rgba(43,108,176,0.3);
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, #3182ce 0%, #2b6cb0 100%);
        border-color: rgba(99,179,237,0.7);
        transform: translateY(-1px);
        box-shadow: 0 6px 16px rgba(43,108,176,0.45);
    }

    /* Download button */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #276749 0%, #1c4532 100%);
        color: #9ae6b4;
        border: 1px solid rgba(104,211,145,0.35);
        border-radius: 10px;
        font-weight: 600;
        font-size: 0.95rem;
        padding: 0.6rem 1.5rem;
        transition: all 0.2s ease;
        width: 100%;
    }
    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #2f855a 0%, #276749 100%);
        border-color: rgba(104,211,145,0.6);
        transform: translateY(-1px);
    }

    /* Dataframe */
    .stDataFrame {
        border-radius: 12px;
        overflow: hidden;
        border: 1px solid rgba(99,179,237,0.15) !important;
    }

    /* Info / warning boxes */
    .stAlert { border-radius: 10px !important; border: none !important; }
    div[data-testid="stInfo"] {
        background: rgba(26, 58, 92, 0.5) !important;
        border-left: 4px solid #63b3ed !important;
    }
    div[data-testid="stSuccess"] {
        background: rgba(27, 67, 50, 0.5) !important;
        border-left: 4px solid #68d391 !important;
    }
    div[data-testid="stWarning"] {
        background: rgba(60, 45, 10, 0.5) !important;
        border-left: 4px solid #f6e05e !important;
    }

    /* Footer */
    .footer-text {
        text-align: center;
        color: #4a6280;
        font-size: 0.78rem;
        margin-top: 3rem;
        padding-top: 1.5rem;
        border-top: 1px solid rgba(99,179,237,0.08);
    }

    /* Expander */
    details {
        background: rgba(15, 30, 50, 0.6) !important;
        border: 1px solid rgba(99,179,237,0.15) !important;
        border-radius: 10px !important;
    }
    summary { color: #90cdf4 !important; font-weight: 600 !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# OCR Engine (cached so it loads only once per session)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def load_ocr_engine():
    return RapidOCR()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
MIN_TEXT_CHARS = 80
AMT_RE_GLOBAL  = re.compile(r"(?:Rs\.?\s*)?(\d{1,3}(?:,\d{2,3})*\.\d{2})")

# =============================================================================
# BANK PROFILES & FORMAT REGISTRY
# =============================================================================
# Two canonical schemas sourced directly from real PDF headers:
#
#  SCHEMA A – Reference: 00007.pdf (Velocity Banking Corporation, etc.)
#  Camelot header row (8 cols):
#    Txn Date | Value Date | Cheque No. | Description | Branch\nCode | Debit | Credit | Balance
#
#  SCHEMA B – Reference: 00018.pdf (INDUSBANK LIMITED, UNITY NATIONAL BANK, etc.)
#  Camelot header row (9 cols):
#    No. | Transaction\nID | Value Date | Txn Posted Date | Cheque No. | Description | Cr/Dr | Transaction\nAmount | Available\nBalance
#
# New banks → add a BANK_PROFILES entry + confirm which schema they use.
# =============================================================================

BANK_PROFILES = [
    # ── Schema A banks ──────────────────────────────────────────────────────
    {
        "bank_id":    "velocity",
        "bank_name":  "Velocity Banking Corporation",
        "keywords":   ["velocity banking", "velo0"],
        "layout_type": "schema_a",
    },
    {
        "bank_id":    "prosperity",
        "bank_name":  "Prosperity Bank Limited",
        "keywords":   ["prosperity bank", "pros0"],
        "layout_type": "schema_a",
    },
    {
        "bank_id":    "metropolitan",
        "bank_name":  "Metropolitan Bank",
        "keywords":   ["metropolitan bank", "metr0"],
        "layout_type": "schema_a",
    },
    # ── Schema B banks ──────────────────────────────────────────────────────
    {
        "bank_id":    "indusbank",
        "bank_name":  "INDUSBANK LIMITED",
        "keywords":   ["indusbank", "indu0"],
        "layout_type": "schema_b",
    },
    {
        "bank_id":    "unity",
        "bank_name":  "UNITY NATIONAL BANK",
        "keywords":   ["unity national bank", "unit0"],
        "layout_type": "schema_b",
    },
    # ── Add new banks here ───────────────────────────────────────────────────
]

FORMAT_REGISTRY = [
    # ═══════════════════════════════════════════════════════════════════════
    # SCHEMA B  –  Cr/Dr Flag with Txn Posted Date  (checked FIRST)
    # Source: 00018.pdf  (INDUSBANK LIMITED, UNITY NATIONAL BANK)
    #
    # Camelot header (9 columns, exact text):
    #   0: No.  |  1: Transaction\nID  |  2: Value Date  |  3: Txn Posted Date
    #   4: Cheque No.  |  5: Description  |  6: Cr/Dr
    #   7: Transaction\nAmount  |  8: Available\nBalance
    #
    # OCR header row (after Y-grouping, Row 21 in 00005.pdf):
    #   ['No.', 'ID', 'Value Date', 'Txn Posted Date', 'Cheque No.',
    #    'Description', 'Cr/Dr', 'Amount', 'Balance']
    # OCR splits 'Transaction\nID'   → 'Transaction' (prev row) + 'ID'
    # OCR splits 'Transaction\nAmount' → 'Transaction' (prev row) + 'Amount'
    # OCR splits 'Available\nBalance'  → 'Available' (prev row) + 'Balance'
    # Key discriminators present in BOTH Camelot & OCR: 'cr/dr', 'txn posted date'
    # ═══════════════════════════════════════════════════════════════════════
    {
        "id":   "schema_b",
        "name": "Schema B – Cr/Dr Flag + Txn Posted Date [ref: 00018.pdf]",

        # Discriminators: 'cr/dr' is unique to Schema B (never in Schema A).
        # 'txn posted date' provides a second lock – present in both Camelot & OCR.
        "required_groups": [
            ["cr/dr"],             # col 6 – in both Camelot & OCR header row
            ["txn posted date"],   # col 3 – in both Camelot & OCR header row
        ],

        # col_roles handle BOTH Camelot multi-line text and OCR short names
        "col_roles": {
            "serial":      ["no."],
            # Camelot: 'transaction\nid'  |  OCR: 'id'
            "txn_id":      ["transaction\nid", "transaction id", "id"],
            "val_date":    ["value date"],
            "posted_date": ["txn posted date"],
            "cheque":      ["cheque no"],
            "description": ["description"],
            "cr_dr_flag":  ["cr/dr"],
            # Camelot: 'transaction\namount' | OCR: 'amount'
            "amount":      ["transaction\namount", "transaction amount", "amount"],
            # Camelot: 'available\nbalance'  | OCR: 'balance'
            "balance":     ["available\nbalance", "available balance", "balance"],
        },

        "cr_dr_mode": True,
    },

    # ═══════════════════════════════════════════════════════════════════════
    # SCHEMA A  –  Standard Debit / Credit columns  (checked SECOND)
    # Source: 00007.pdf  (Velocity Banking Corporation)
    # Exact Camelot header (8 columns, index 0-7):
    #   0: Txn Date  |  1: Value Date  |  2: Cheque No.
    #   3: Description  |  4: Branch\nCode  |  5: Debit  |  6: Credit  |  7: Balance
    # OCR header row:  same labels (single-line, no split needed)
    # ═══════════════════════════════════════════════════════════════════════
    {
        "id":   "schema_a",
        "name": "Schema A – Standard (Debit / Credit) [ref: 00007.pdf]",

        # ALL groups must match at least one header cell
        "required_groups": [
            ["txn date"],   # col 0 – exact; absent in Schema B
            ["debit"],      # col 5 – exact; absent in Schema B
            ["credit"],     # col 6 – exact; absent in Schema B
            ["balance"],    # col 7 – exact
        ],

        "col_roles": {
            "txn_date":    ["txn date"],
            "val_date":    ["value date"],
            "cheque":      ["cheque no"],
            "description": ["description"],
            "branch":      ["branch"],
            "debit":       ["debit"],
            "credit":      ["credit"],
            "balance":     ["balance"],
        },

        "cr_dr_mode": False,
    },
]



def _clean_amt(s: str, amt_re) -> str:
    """Extract first amount string from a cell, stripping Rs. prefix."""
    m = amt_re.search(str(s))
    return m.group(1) if m else ""


def _detect_format(rows_of_cells: list) -> tuple:
    """
    Given the first few rows (each = list of lowercase string cells),
    try to match an entry in FORMAT_REGISTRY.

    Returns (fmt_def, header_row_idx, col_map) or (None, None, {}).
    col_map: {role -> column_index}
    """
    for row_idx, cells in enumerate(rows_of_cells[:8]):     # scan first 8 rows
        cells_lower = [str(c).strip().lower() for c in cells]

        for fmt in FORMAT_REGISTRY:
            # Every required_group must have at least one keyword matching some cell
            all_found = True
            for group in fmt["required_groups"]:
                if not any(
                    any(kw in cell for kw in group)
                    for cell in cells_lower
                ):
                    all_found = False
                    break

            if not all_found:
                continue

            # Build col_map: first cell whose text contains any keyword for this role
            col_map = {}
            for role, keywords in fmt["col_roles"].items():
                for col_i, cell in enumerate(cells_lower):
                    if any(kw in cell for kw in keywords):
                        col_map.setdefault(role, col_i)
                        break

            return fmt, row_idx, col_map

    return None, None, {}


def _get_cell(cells: list, col_map: dict, role: str, default: str = "") -> str:
    """Safely retrieve a cell value by role name."""
    idx = col_map.get(role)
    if idx is None or idx >= len(cells):
        return default
    return cells[idx]


def _extract_row_standard(cells, col_map, amt_pattern, date_re, val_date_re):
    """
    Format A – separate Debit / Credit columns.
    Returns a row dict (with _direct=True) or None.
    """
    txn_date_raw = _get_cell(cells, col_map, "txn_date")
    date_m = date_re.search(txn_date_raw)
    if not date_m:
        return None
    txn_date = date_m.group(1)

    val_date_raw = _get_cell(cells, col_map, "val_date")
    vdm = date_re.search(val_date_raw)
    val_date = vdm.group(1) if vdm else ""

    desc        = re.sub(r"\s+", " ", _get_cell(cells, col_map, "description")).strip()
    cheque      = _get_cell(cells, col_map, "cheque").strip()
    branch      = _get_cell(cells, col_map, "branch").strip()
    debit_amt   = _clean_amt(_get_cell(cells, col_map, "debit"),   amt_pattern)
    credit_amt  = _clean_amt(_get_cell(cells, col_map, "credit"),  amt_pattern)
    balance_amt = _clean_amt(_get_cell(cells, col_map, "balance"), amt_pattern)

    if not balance_amt and not debit_amt and not credit_amt:
        return None

    return {
        "Txn Date":    txn_date,
        "Value Date":  val_date,
        "Cheque No":   cheque,
        "Description": desc,
        "Branch Code": branch,
        "Debit":       f"Rs. {debit_amt}"   if debit_amt   else "",
        "Credit":      f"Rs. {credit_amt}"  if credit_amt  else "",
        "Balance":     f"Rs. {balance_amt}" if balance_amt else "",
        "_direct":     True,
        "Raw_Amounts": [],
    }


def _extract_row_crdr(cells, col_map, amt_pattern, date_re, val_date_re):
    """
    Format B – includes Txn Posted Date, Txn ID, Serial No, Cheque, Cr/Dr Flag, Amount, Balance.
    Returns a row dict (with _direct=True) or None.
    """
    serial = _get_cell(cells, col_map, "serial").split("\n")[0].strip()
    txn_id = _get_cell(cells, col_map, "txn_id").strip()

    val_date_raw = _get_cell(cells, col_map, "val_date")
    vdm = date_re.search(val_date_raw)
    val_date = vdm.group(1) if vdm else val_date_raw.replace("\n", " ").strip()

    posted_date_raw = _get_cell(cells, col_map, "posted_date") or _get_cell(cells, col_map, "txn_date")
    posted_date = posted_date_raw.replace("\n", " ").strip()

    desc      = re.sub(r"\s+", " ", _get_cell(cells, col_map, "description")).strip()
    cheque    = _get_cell(cells, col_map, "cheque").strip()
    flag_raw  = _get_cell(cells, col_map, "cr_dr_flag").strip().upper()
    amt_raw   = _get_cell(cells, col_map, "amount")
    bal_raw   = _get_cell(cells, col_map, "balance")

    amount_amt  = _clean_amt(amt_raw, amt_pattern)
    balance_amt = _clean_amt(bal_raw, amt_pattern)

    if not amount_amt:
        return None

    is_credit = "CR" in flag_raw
    is_debit  = "DR" in flag_raw
    if not is_credit and not is_debit:
        return None

    return {
        "No":                 serial,
        "Transaction ID":     txn_id,
        "Value Date":         val_date,
        "Txn Posted Date":    posted_date,
        "Cheque No":          cheque,
        "Description":        desc,
        "Cr/Dr":              "CR" if is_credit else "DR",
        "Transaction Amount": f"Rs. {amount_amt}",
        "Debit":              f"Rs. {amount_amt}" if is_debit  else "",
        "Credit":             f"Rs. {amount_amt}" if is_credit else "",
        "Balance":            f"Rs. {balance_amt}" if balance_amt else "",
        "_direct":            True,
        "Raw_Amounts":        [],
    }


# Dispatch table: schema id → row extractor function
_ROW_EXTRACTORS = {
    "schema_a": _extract_row_standard,   # 00007.pdf layout (Debit / Credit columns)
    "schema_b": _extract_row_crdr,       # 00018.pdf layout (Cr/Dr Flag + Txn Posted Date)
}

# Comprehensive metadata regex patterns
_META_PATTERNS = {
    "Customer ID":       r"(?:customer\s*(?:id|no|cif|number))[.:\s]+([A-Z0-9]{5,20})",
    "Branch Name":       r"(?:branch\s*(?:name)?)[.:\s]+([A-Za-z][A-Za-z .\-]{2,50})",
    "IFSC Code":         r"(?:ifsc\s*(?:code)?)[.:\s]+([A-Z]{4}0[A-Z0-9]{6})",
    "MICR Code":         r"(?:micr\s*(?:code)?)[.:\s]+(\d{9})",
    "Account Number":    r"(?:account\s*number)[.:\s]+([0-9]{8,20})",
    "Account Type":      r"(?:account\s*type|product\s*name)[.:\s]+([A-Za-z0-9 \-_]{3,40})",
    "Account Currency":  r"(?:account\s*currency|currency)[.:\s]+([A-Za-z]{3})",
    "Interest Rate":     r"(?:interest\s*rate)[.:\s]+([0-9.]+\s*%\s*p\.?a\.?)",
    "Statement Period":  r"(?:searched\s*by|period|statement\s*(?:period|from))[.:\s]+(?:from\s*)?(\d{1,2}\s+[A-Za-z]{3}\s+\d{4}\s*(?:to|-)\s*\d{1,2}\s+[A-Za-z]{3}\s+\d{4}|\d{2}[/-]\d{2}[/-]\d{4}\s*(?:to|-)\s*\d{2}[/-]\d{2}[/-]\d{4})",
    "Opening Balance":   r"(?:opening\s*balance)[.:\s]+(?:Rs\.?\s*)?(\d[\d,]*\.\d{2})",
    "Closing Balance":   r"(?:closing\s*balance)[.:\s]+(?:Rs\.?\s*)?(\d[\d,]*\.\d{2})",
    "Email":             r"(?:email\s*(?:id)?)[.:\s]+([\w.\-+]+@[\w.\-]+\.[a-z]{2,})",
    "Mobile":            r"(?:mobile|phone|contact)[.:\s]+(\+?\d[\d \-]{8,14})",
    "PAN":               r"\b([A-Z]{5}\d{4}[A-Z])\b",
}


def _page_is_scanned(page) -> bool:
    """Return True if the PyMuPDF page has little/no selectable text (i.e. scanned)."""
    return len(page.get_text("text").strip()) < MIN_TEXT_CHARS


def _clean_amt(s: str, amt_re) -> str:
    """Extract first amount string from a cell, stripping Rs. prefix."""
    m = amt_re.search(str(s))
    return m.group(1) if m else ""


def _detect_bank_name(text: str) -> str:
    """Detect Bank Name using bank profiles and keyword heuristics."""
    text_lower = text.lower()
    for prof in BANK_PROFILES:
        if any(kw in text_lower for kw in prof["keywords"]):
            return prof["bank_name"]
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    return lines[0] if lines else "Bank Statement"


def _extract_full_metadata(doc, page_types: dict, dpi: int = 200) -> dict:
    """
    Extract comprehensive account metadata, client name, and client address
    from Page 1 (using native text for digital PDFs, or RapidOCR for scanned PDFs).
    """
    metadata = {}
    full_text = ""

    # Check if page 1 is scanned
    is_p1_scanned = page_types.get(1, {}).get("layer") == "scanned" if isinstance(page_types.get(1), dict) else page_types.get(1) == "scanned"

    if is_p1_scanned:
        try:
            ppocr_engine = load_ocr_engine()
            pix = doc[0].get_pixmap(dpi=dpi)
            img = cv2.imdecode(np.frombuffer(pix.tobytes("png"), np.uint8), cv2.IMREAD_COLOR)
            res = ppocr_engine(img)
            if res and res.txts:
                full_text = "\n".join(res.txts)
        except Exception:
            full_text = ""
    else:
        for pg in range(min(2, len(doc))):
            try:
                full_text += doc[pg].get_text("text") + "\n"
            except Exception:
                pass

    if not full_text:
        return metadata

    # 1. Bank Name
    metadata["Bank Name"] = _detect_bank_name(full_text)

    # 2. Client Name & Address extraction
    lines = [l.strip() for l in full_text.splitlines() if l.strip()]
    start_idx = None
    for idx, l in enumerate(lines[:10]):
        if "statement" in l.lower() and idx < 4:
            start_idx = idx + 1
            break

    if start_idx is not None and start_idx < len(lines):
        client_lines = []
        for l in lines[start_idx:]:
            if any(term in l.lower() for term in [
                "account statement as of", "account holder", "customer id", "branch", "micr", "ifsc", "account number"
            ]):
                break
            client_lines.append(l)

        if client_lines:
            metadata["Account Holder"] = client_lines[0]
            if len(client_lines) > 1:
                metadata["Client Address"] = ", ".join(client_lines[1:])

    # 3. Regex Patterns for metadata table
    for field, pattern in _META_PATTERNS.items():
        m = re.search(pattern, full_text, re.IGNORECASE)
        if m:
            metadata[field] = m.group(1).strip()

    return metadata


def _apply_math_engine(df: pd.DataFrame) -> pd.DataFrame:
    """
    Classify Debit/Credit using balance-change math.
    Rows already classified by Camelot (flag _direct=True) are passed through.
    """
    debits, credits, balances = [], [], []
    prev_balance = None

    for _, row in df.iterrows():
        # Directly classified by Camelot column headers — trust those values
        if row.get("_direct"):
            debits.append(row.get("Debit", ""))
            credits.append(row.get("Credit", ""))
            balances.append(row.get("Balance", ""))
            try:
                b = row.get("Balance", "")
                prev_balance = float(b.replace("Rs.", "").replace(",", "").strip()) if b else prev_balance
            except Exception:
                pass
            continue

        raw_amts = row.get("Raw_Amounts", [])
        desc     = str(row.get("Description", "")).lower()

        if len(raw_amts) >= 2:
            bal_num = float(raw_amts[1].replace(",", ""))
            txn_str = f"Rs. {raw_amts[0]}"
            bal_str = f"Rs. {raw_amts[1]}"

            if prev_balance is not None:
                diff = round(bal_num - prev_balance, 2)
                if diff > 0:
                    credits.append(txn_str); debits.append("")
                else:
                    debits.append(txn_str); credits.append("")
            else:
                if any(kw in desc for kw in ["by clg", "cr", "salary", "deposit", "upi/cr"]):
                    credits.append(txn_str); debits.append("")
                else:
                    debits.append(txn_str); credits.append("")

            balances.append(bal_str)
            prev_balance = bal_num

        elif len(raw_amts) == 1:
            debits.append(""); credits.append("")
            balances.append(f"Rs. {raw_amts[0]}")
            try:
                prev_balance = float(raw_amts[0].replace(",", ""))
            except Exception:
                pass
        else:
            debits.append(""); credits.append(""); balances.append("")

    df["Debit"]   = debits
    df["Credit"]  = credits
    df["Balance"] = balances

    # Drop internal columns
    drop_cols = [c for c in ["Raw_Amounts", "_direct"] if c in df.columns]
    if drop_cols:
        df.drop(columns=drop_cols, inplace=True)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Camelot extractor  (digital / text-layer PDFs)
# ─────────────────────────────────────────────────────────────────────────────
def _extract_with_camelot(tmp_path: str, page_num: int,
                          date_re, val_date_re, amt_pattern) -> list:
    """
    Run Camelot on one digital page with FORMAT_REGISTRY detection.

    For each table:
      1. Feed first 8 rows to _detect_format() → returns format + header row + col_map
      2. If format found → use the registered row-extractor (exact column positions).
      3. If no format found → regex/math-engine fallback.
    """
    parsed_rows = []
    chq_re    = re.compile(r"\b(\d{6})\b")
    branch_re = re.compile(r"\b(\d{4})\b")

    try:
        tables = camelot.read_pdf(tmp_path, pages=str(page_num), flavor="lattice")
        if not tables or tables.n == 0:
            tables = camelot.read_pdf(tmp_path, pages=str(page_num), flavor="stream")
    except Exception:
        return parsed_rows

    for table in tables:
        df_raw = table.df
        if df_raw.empty:
            continue

        # Build list-of-list-of-lowercase-strings for format detection
        sample_rows = [
            [str(c).strip().lower() for c in df_raw.iloc[i]]
            for i in range(min(8, len(df_raw)))
        ]
        fmt, header_row_idx, col_map = _detect_format(sample_rows)

        # ── Path A: known format detected → use registered row extractor ────────
        if fmt is not None:
            extractor = _ROW_EXTRACTORS.get(fmt["id"])
            if extractor:
                data_rows = df_raw.iloc[header_row_idx + 1:]
                for _, row in data_rows.iterrows():
                    cells = [str(c).strip() for c in row]
                    result = extractor(cells, col_map, amt_pattern, date_re, val_date_re)
                    if result:
                        parsed_rows.append(result)
                continue   # done with this table

        # ── Path B: unknown format → regex fallback (math engine will classify) ──
        for _, row in df_raw.iterrows():
            line_str = " | ".join(str(c).strip() for c in row if str(c).strip())

            date_m = date_re.search(line_str)
            if not date_m or "statement as of" in line_str.lower():
                continue

            txn_date   = date_m.group(1)
            val_date_m = val_date_re.search(line_str)
            val_date   = val_date_m.group(1) if val_date_m else ""
            amounts    = amt_pattern.findall(line_str)
            if not amounts:
                continue

            working_str = date_re.sub("", line_str, count=1)
            if val_date:
                working_str = val_date_re.sub("", working_str, count=1)

            chq_no = ""
            chq_m  = chq_re.search(working_str)
            if chq_m and working_str.find(chq_m.group(1)) < working_str.find(amounts[0]):
                chq_no      = chq_m.group(1)
                working_str = working_str.replace(chq_no, "", 1)

            branch_code = ""
            branch_m    = branch_re.search(working_str)
            if branch_m:
                branch_code = branch_m.group(1)
                working_str = working_str.replace(branch_code, "", 1)

            working_str = amt_pattern.sub("", working_str)
            description = re.sub(r"[|\s]+", " ", working_str).strip()

            parsed_rows.append({
                "Txn Date":    txn_date,
                "Value Date":  val_date,
                "Cheque No":   chq_no,
                "Description": description,
                "Branch Code": branch_code,
                "_direct":     False,
                "Raw_Amounts": amounts,
            })
    return parsed_rows


# ─────────────────────────────────────────────────────────────────────────────
# OCR extractor  (scanned / image-based PDFs)
# ─────────────────────────────────────────────────────────────────────────────
def _extract_with_ocr(page, page_num: int, dpi: int,
                      date_re, val_date_re, chq_re, branch_re, amt_pattern) -> list:
    """
    Run RapidOCR on one scanned page with FORMAT_REGISTRY detection.

    Flow:
      1. OCR all tokens, group by Y-position into rows.
      2. Feed first 8 rows (text cells) to _detect_format().
      3. If format found:
           - Record the X-centre of each detected column from the header row.
           - For each subsequent row, assign tokens to the nearest column bucket.
           - Call the registered row-extractor with the bucketed cells.
      4. If no format found (or extractor fails): regex/math-engine fallback.
    """
    ppocr_engine = load_ocr_engine()
    parsed_rows  = []

    pix = page.get_pixmap(dpi=dpi)
    img = cv2.imdecode(
        np.frombuffer(pix.tobytes("png"), np.uint8), cv2.IMREAD_COLOR
    )

    ocr_out = ppocr_engine(img)
    if ocr_out is None or ocr_out.boxes is None or len(ocr_out.boxes) == 0:
        return parsed_rows

    # Collect items with position
    items = []
    for box, txt, score in zip(ocr_out.boxes, ocr_out.txts, ocr_out.scores):
        min_x = min(p[0] for p in box)
        max_x = max(p[0] for p in box)
        min_y = min(p[1] for p in box)
        items.append({"min_x": min_x, "max_x": max_x,
                      "cx": (min_x + max_x) / 2,
                      "min_y": min_y, "text": txt.strip()})

    items.sort(key=lambda it: it["min_y"])

    # Group into rows by Y proximity
    rows = []           # list of [item, ...]
    curr_row, curr_y = [], None
    for item in items:
        y = item["min_y"]
        if curr_y is None or abs(y - curr_y) < 18:
            curr_row.append(item)
            curr_y = y if curr_y is None else (curr_y + y) / 2
        else:
            curr_row.sort(key=lambda it: it["min_x"])
            rows.append(curr_row)
            curr_row = [item]
            curr_y   = y
    if curr_row:
        curr_row.sort(key=lambda it: it["min_x"])
        rows.append(curr_row)

    # Build text-only sample rows for format detection
    sample_cells = [[it["text"].lower() for it in row] for row in rows[:8]]
    fmt, header_row_idx, col_map = _detect_format(sample_cells)

    # ── Path A: format detected → column-position-aware extraction ────────────
    if fmt is not None and header_row_idx is not None:
        extractor  = _ROW_EXTRACTORS.get(fmt["id"])
        header_row = rows[header_row_idx]

        # Record X-centre of each role’s column from the header tokens
        role_x_centers = {}
        for role, col_idx in col_map.items():
            if col_idx < len(header_row):
                role_x_centers[role] = header_row[col_idx]["cx"]

        all_roles = list(role_x_centers.keys())

        def _bucket_row(row_items):
            """Assign each OCR token to the nearest role bucket by X centre."""
            buckets = {r: [] for r in all_roles}
            for it in row_items:
                if not all_roles:
                    break
                nearest = min(all_roles, key=lambda r: abs(role_x_centers[r] - it["cx"]))
                buckets[nearest].append(it["text"])
            return {r: " ".join(v) for r, v in buckets.items()}

        if extractor:
            for data_row in rows[header_row_idx + 1:]:
                role_cells = _bucket_row(data_row)
                # Build a cells list aligned by col_map index
                n_cols = max(col_map.values()) + 1 if col_map else 0
                cells = [""] * n_cols
                for role, idx in col_map.items():
                    cells[idx] = role_cells.get(role, "")

                result = extractor(cells, col_map, amt_pattern, date_re, val_date_re)
                if result:
                    parsed_rows.append(result)

            if parsed_rows:
                return parsed_rows   # done – skip fallback

    # ── Path B: no format / extractor failed → regex + math-engine fallback ──
    for r in rows:
        line_str = " | ".join(it["text"] for it in r)

        date_m = date_re.search(line_str)
        if not date_m or "statement as of" in line_str.lower():
            continue

        txn_date   = date_m.group(1)
        val_date_m = val_date_re.search(line_str)
        val_date   = val_date_m.group(1) if val_date_m else ""
        amounts    = amt_pattern.findall(line_str)
        if not amounts:
            continue

        working_str = line_str
        working_str = date_re.sub("", working_str, count=1)
        if val_date:
            working_str = val_date_re.sub("", working_str, count=1)

        chq_no = ""
        chq_m  = chq_re.search(working_str)
        if chq_m and (not amounts or working_str.find(chq_m.group(1)) < working_str.find(amounts[0])):
            chq_no      = chq_m.group(1)
            working_str = working_str.replace(chq_no, "", 1)

        branch_code = ""
        branch_m    = branch_re.search(working_str)
        if branch_m:
            branch_code = branch_m.group(1)
            working_str = working_str.replace(branch_code, "", 1)

        working_str = amt_pattern.sub("", working_str)
        description = re.sub(r"[|\s]+", " ", working_str).strip()

        parsed_rows.append({
            "Txn Date":    txn_date,
            "Value Date":  val_date,
            "Cheque No":   chq_no,
            "Description": description,
            "Branch Code": branch_code,
            "_direct":     False,
            "Raw_Amounts": amounts,
        })
    return parsed_rows


# ─────────────────────────────────────────────────────────────────────────────
# Main Extraction Orchestrator  (Hybrid: Camelot for digital, OCR for scanned)
# ─────────────────────────────────────────────────────────────────────────────
def extract_bank_statement(pdf_bytes: bytes, dpi: int = 200, progress_cb=None):
    """
    Hybrid extractor:
      - Digital pages  → Camelot (column-header-aware)
      - Scanned pages  → RapidOCR (image-based OCR)
    Returns (df, metadata_dict, page_types_dict).
    """
    date_re     = re.compile(r"(\d{2}[-/]\d{2}[-/]\d{4})")
    val_date_re = re.compile(r"(\d{2}\s+[A-Za-z]{3}\s+\d{4})")
    chq_re      = re.compile(r"\b(\d{6})\b")
    branch_re   = re.compile(r"\b(\d{4})\b")
    amt_pattern = re.compile(r"(?:Rs\.?\s*)?(\d{1,3}(?:,\d{2,3})*\.\d{2})")

    parsed_rows = []
    page_types  = {}   # page_num -> {"layer": "scanned"|"digital", "format": str}

    doc         = pymupdf.open(stream=pdf_bytes, filetype="pdf")
    total_pages = len(doc)

    # Write PDF to a temp file so Camelot (which needs a path) can read it
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    tmp_file.write(pdf_bytes)
    tmp_file.close()
    tmp_path = tmp_file.name

    try:
        for page_idx in range(total_pages):
            page_num = page_idx + 1

            if progress_cb:
                progress_cb(page_idx, total_pages)

            page    = doc[page_idx]
            scanned = _page_is_scanned(page)
            layer   = "scanned" if scanned else "digital"

            if scanned:
                rows = _extract_with_ocr(
                    page, page_num, dpi,
                    date_re, val_date_re, chq_re, branch_re, amt_pattern
                )
            else:
                rows = _extract_with_camelot(
                    tmp_path, page_num,
                    date_re, val_date_re, amt_pattern
                )

            # Determine which format was detected (if any _direct row exists)
            detected_fmt = "Fallback (Math Engine)"
            for r in rows:
                if r.get("_direct"):
                    # Infer from presence of cr_dr_flag or standard columns
                    has_sep = r.get("Debit") != "" or r.get("Credit") != ""
                    # We know if Format B was used there's no Branch Code
                    detected_fmt = "Format B – Cr/Dr Flag" if not r.get("Branch Code") and r.get("_direct") else "Format A – Standard"
                    break

            page_types[page_num] = {"layer": layer, "format": detected_fmt}
            parsed_rows.extend(rows)

        # Extract full metadata from native text (digital pages preferred)
        metadata = _extract_full_metadata(doc, page_types, dpi=dpi)

    finally:
        doc.close()
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    if not parsed_rows:
        return pd.DataFrame(), metadata, page_types

    df = pd.DataFrame(parsed_rows)
    df = _apply_math_engine(df)

    return df, metadata, page_types


# ─────────────────────────────────────────────────────────────────────────────
# Styled Excel Builder
# ─────────────────────────────────────────────────────────────────────────────
def build_styled_excel(df: pd.DataFrame, metadata: dict) -> bytes:
    """
    Build a styled multi-section Excel workbook:
      1. Main 'Transactions' Sheet:
         - Top Banner: Bank Name & Statement Title
         - Metadata Summary Cards: Party Name, Address, Account No, Cust ID,
           Bank Name, Branch Name, IFSC, MICR, Statement Period, Balances
         - Transaction Ledger: Full table with Debit/Credit color coding,
           proper column alignment, width formatting, and frozen panes.
      2. Secondary 'Account_Metadata' Sheet:
         - Complete tabular list of all extracted key-value metadata fields.
    """
    import openpyxl

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Total columns to span
    num_cols = max(len(df.columns) if not df.empty else 8, 9)
    last_col_letter = get_column_letter(num_cols)

    # 1. Main Title Banner (Row 1)
    bank_title = metadata.get("Bank Name", "BANK STATEMENT").upper()
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = f"🏦  {bank_title}  –  ACCOUNT STATEMENT"
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.fill = PatternFill("solid", fgColor="1A3A5C")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row 2: subtle divider spacer
    ws.row_dimensions[2].height = 6

    # 2. Top Metadata Summary Box (Rows 3 to N)
    lbl_font = Font(bold=True, color="63B3ED", size=9)
    val_font = Font(bold=False, color="FFFFFF", size=9)
    box_fill = PatternFill("solid", fgColor="0D1B2A")
    hdr_box_fill = PatternFill("solid", fgColor="16283D")

    thin_side = Side(style="thin", color="2B4F72")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    mid_col_idx = max(4, num_cols // 2)
    mid_col_letter = get_column_letter(mid_col_idx)
    next_col_letter = get_column_letter(mid_col_idx + 1)

    # Row 3: Section Headers
    ws.merge_cells(f"A3:{mid_col_letter}3")
    sec1 = ws["A3"]
    sec1.value = "👤  PARTY & ACCOUNT DETAILS"
    sec1.font = Font(bold=True, color="90CDF4", size=10)
    sec1.fill = hdr_box_fill
    sec1.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells(f"{next_col_letter}3:{last_col_letter}3")
    sec2 = ws[f"{next_col_letter}3"]
    sec2.value = "🏦  BANK & STATEMENT DETAILS"
    sec2.font = Font(bold=True, color="90CDF4", size=10)
    sec2.fill = hdr_box_fill
    sec2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 22

    # Metadata items
    party_address = metadata.get("Client Address") or metadata.get("Address", "—")
    op_bal = metadata.get("Opening Balance", "")
    cl_bal = metadata.get("Closing Balance", "")

    left_meta = [
        ("Account Holder / Party", metadata.get("Account Holder", "—")),
        ("Party Address",          party_address),
        ("Account Number",         metadata.get("Account Number", "—")),
        ("Customer ID",            metadata.get("Customer ID", "—")),
        ("Account Type",           metadata.get("Account Type", "—")),
    ]
    if metadata.get("Nominee"):
        left_meta.append(("Nominee", metadata.get("Nominee")))
    if metadata.get("PAN"):
        left_meta.append(("PAN", metadata.get("PAN")))

    right_meta = [
        ("Bank Name",        metadata.get("Bank Name", "—")),
        ("Branch Name",      metadata.get("Branch Name", "—")),
        ("IFSC Code",        metadata.get("IFSC Code", "—")),
        ("MICR Code",        metadata.get("MICR Code", "—")),
        ("Statement Period", metadata.get("Statement Period", "—")),
        ("Opening Balance",  f"Rs. {op_bal}" if op_bal else "—"),
        ("Closing Balance",  f"Rs. {cl_bal}" if cl_bal else "—"),
    ]
    if metadata.get("Interest Rate"):
        right_meta.append(("Interest Rate", metadata.get("Interest Rate")))

    max_meta_rows = max(len(left_meta), len(right_meta))

    for idx in range(max_meta_rows):
        r = 4 + idx
        ws.row_dimensions[r].height = 20

        # Left side: Column A (Label), Columns B..mid_col_idx (Value)
        if idx < len(left_meta):
            lbl, val = left_meta[idx]
            lbl_cell = ws[f"A{r}"]
            lbl_cell.value = lbl
            lbl_cell.font = lbl_font
            lbl_cell.fill = box_fill
            lbl_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            val_cell = ws[f"B{r}"]
            val_cell.value = val
            val_cell.font = val_font
            val_cell.fill = box_fill
            val_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            if mid_col_idx > 2:
                ws.merge_cells(f"B{r}:{mid_col_letter}{r}")
            for c_i in range(1, mid_col_idx + 1):
                ws.cell(row=r, column=c_i).border = Border(
                    left=Side(style="thin", color="16283D"),
                    right=Side(style="thin", color="16283D"),
                    top=Side(style="thin", color="16283D"),
                    bottom=Side(style="thin", color="16283D"),
                )
        else:
            for c_i in range(1, mid_col_idx + 1):
                c_cell = ws.cell(row=r, column=c_i)
                c_cell.fill = box_fill

        # Right side: Column next_col_letter (Label), remaining cols (Value)
        if idx < len(right_meta):
            lbl, val = right_meta[idx]
            lbl_cell = ws[f"{next_col_letter}{r}"]
            lbl_cell.value = lbl
            lbl_cell.font = lbl_font
            lbl_cell.fill = box_fill
            lbl_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            val_col_letter = get_column_letter(mid_col_idx + 2)
            val_cell = ws[f"{val_col_letter}{r}"]
            val_cell.value = val
            val_cell.font = val_font
            val_cell.fill = box_fill
            val_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            if num_cols > mid_col_idx + 2:
                ws.merge_cells(f"{val_col_letter}{r}:{last_col_letter}{r}")
            for c_i in range(mid_col_idx + 1, num_cols + 1):
                ws.cell(row=r, column=c_i).border = Border(
                    left=Side(style="thin", color="16283D"),
                    right=Side(style="thin", color="16283D"),
                    top=Side(style="thin", color="16283D"),
                    bottom=Side(style="thin", color="16283D"),
                )
        else:
            for c_i in range(mid_col_idx + 1, num_cols + 1):
                c_cell = ws.cell(row=r, column=c_i)
                c_cell.fill = box_fill

    table_start_row = 4 + max_meta_rows + 2

    # Row before table: Section Header
    ws.merge_cells(f"A{table_start_row-1}:{last_col_letter}{table_start_row-1}")
    tbl_title = ws[f"A{table_start_row-1}"]
    tbl_title.value = "📊  TRANSACTION LEDGER"
    tbl_title.font = Font(bold=True, color="63B3ED", size=11)
    tbl_title.fill = PatternFill("solid", fgColor="0D1B2A")
    tbl_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[table_start_row - 1].height = 26

    # Write Table Column Headers
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="90CDF4", size=10)

    if not df.empty:
        col_names = list(df.columns)
        for c_idx, col_name in enumerate(col_names, 1):
            cell = ws.cell(row=table_start_row, column=c_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[table_start_row].height = 24

        # Write Data Rows
        debit_fill  = PatternFill("solid", fgColor="2D1515")
        credit_fill = PatternFill("solid", fgColor="0F2F1A")
        even_fill   = PatternFill("solid", fgColor="111E2D")
        odd_fill    = PatternFill("solid", fgColor="0D1826")
        data_font   = Font(color="D1E8F5", size=9)

        col_indices = {col_name: i + 1 for i, col_name in enumerate(col_names)}

        for row_offset, (_, row_data) in enumerate(df.iterrows(), start=1):
            curr_r = table_start_row + row_offset
            ws.row_dimensions[curr_r].height = 19

            debit_val  = str(row_data.get("Debit",  "") or "").strip()
            credit_val = str(row_data.get("Credit", "") or "").strip()

            if debit_val:
                bg = debit_fill
            elif credit_val:
                bg = credit_fill
            else:
                bg = even_fill if row_offset % 2 == 0 else odd_fill

            for c_idx, col_name in enumerate(col_names, 1):
                val = row_data[col_name]
                cell = ws.cell(row=curr_r, column=c_idx, value=val)
                cell.fill = bg
                cell.font = data_font
                cell.border = Border(
                    left=Side(style="thin", color="1A2D40"),
                    right=Side(style="thin", color="1A2D40"),
                )

                # Alignment rules
                if col_name in ["Debit", "Credit", "Balance", "Transaction Amount"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in ["No", "Cheque No", "Branch Code", "Cr/Dr", "Value Date", "Transaction Direction", "Transaction Mode"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Column Widths
        col_widths = {
            "No": 8, "Transaction ID": 18, "Txn Posted Date": 24,
            "Txn Date": 16, "Value Date": 16, "Cheque No": 14,
            "Description": 48, "Branch Code": 14, "Cr/Dr": 10,
            "Transaction Amount": 20, "Debit": 18, "Credit": 18, "Balance": 20,
            "Verification Status": 22,
            "Source File": 28,
            "Transaction Mode": 20,
            "Transaction Direction": 18,
            "Transaction Purpose": 25,
            "Bank / Institution": 25,
            "Party / Counterparty": 30,
        }
        for col_name, width in col_widths.items():
            if col_name in col_indices:
                ws.column_dimensions[get_column_letter(col_indices[col_name])].width = width

        ws.freeze_panes = f"A{table_start_row+1}"

    # 3. Secondary 'Account_Metadata' Sheet
    if metadata:
        ws_meta = wb.create_sheet(title="Account_Metadata")
        ws_meta.append(["Metadata Field", "Extracted Value"])
        for k, v in metadata.items():
            ws_meta.append([k, str(v)])

        for cell in ws_meta[1]:
            cell.fill = PatternFill("solid", fgColor="1A3A5C")
            cell.font = Font(bold=True, color="90CDF4", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_meta.row_dimensions[1].height = 24

        meta_border = Border(
            left=Side(style="thin", color="2B4F72"),
            right=Side(style="thin", color="2B4F72"),
            top=Side(style="thin", color="2B4F72"),
            bottom=Side(style="thin", color="2B4F72"),
        )
        for row in ws_meta.iter_rows(min_row=2, max_row=ws_meta.max_row):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="0D1B2A")
                cell.font = Font(color="D1E8F5", size=10)
                cell.border = meta_border

        ws_meta.column_dimensions["A"].width = 28
        ws_meta.column_dimensions["B"].width = 55

    wb.save(output)
    output.seek(0)
    return output.read()


# ─────────────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        "<div style='text-align:center;padding:1rem 0 0.5rem 0;'>"
        "<div style='font-size:2rem;'>🏦</div>"
        "<div style='font-weight:800;font-size:1.1rem;color:#90cdf4;letter-spacing:0.06em;'>BankLens AI</div>"
        "<div style='font-size:0.72rem;color:#4a6e8a;margin-top:0.2rem;'>Transaction Intelligence Platform</div>"
        "</div>",
        unsafe_allow_html=True,
    )
    st.markdown("---")

    dpi_val = st.slider(
        "📷 OCR Resolution (DPI)",
        min_value=150, max_value=300, value=200, step=25,
        help="Higher DPI = better accuracy on scanned statements but slower."
    )

    st.markdown("---")
    st.markdown("### 🔄 Pipeline Steps")
    st.markdown(
        """
        1. **Upload** bank statement PDF(s)
        2. **Auto-detect** page type (digital / scanned)
        3. **Extract** table rows via Camelot or RapidOCR
        4. **Verify** running balance mathematically
        5. **Clean** transaction descriptions
        6. **Embed** descriptions with Nomic AI
        7. **Classify** 5 domain categories (Mode, Direction, Purpose, Bank, Counterparty)
        8. **Download** the enriched Excel with all 5 columns
        """
    )

    st.markdown("---")
    st.markdown("### 📊 5 Classification Columns")
    for label, vals in [
        ("💳 Mode", "UPI · NEFT · RTGS · IMPS · ATM · Cash · Cheque · NetBanking"),
        ("↕️ Direction", "Debit · Credit"),
        ("🎯 Purpose", "Salary · Rent · Utility · Healthcare · Investment · SaaS…"),
        ("🏛️ Bank", "HDFC · Citi · ICICI · SBI · Axis · Kotak…"),
        ("👤 Counterparty", "Person / Merchant name extracted"),
    ]:
        st.markdown(
            f"<div style='margin-bottom:0.5rem;'>"
            f"<span style='color:#90cdf4;font-weight:700;font-size:0.8rem;'>{label}</span><br>"
            f"<span style='color:#7a9eba;font-size:0.76rem;'>{vals}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )

    st.markdown("---")
    st.markdown(
        "<div style='color:#2d4a60;font-size:0.72rem;text-align:center;'>"
        "BankLens v2.0 · Camelot + RapidOCR + Nomic AI"
        "</div>",
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Classifier Engine Cache
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner="⚡ Loading AI Embedding Model & 5-Class Taxonomies…")
def load_classifier_engine():
    return get_classification_engine()


# ─────────────────────────────────────────────────────────────────────────────
# Hero Banner
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <div class="hero-banner">
        <h1>🏦 BankLens AI</h1>
        <p>Upload any bank statement PDF — get a fully structured, AI-classified Excel with 5 domain categories automatically appended to every transaction</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: Upload
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<div class="step-badge">STEP 1 — Upload PDF Statement(s)</div>', unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "Upload bank statement PDFs (digital or scanned)",
    type=["pdf"],
    accept_multiple_files=True,
    key="pdf_uploader",
    label_visibility="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# Main Pipeline: Extract → Classify → Store → Display → Download
# ─────────────────────────────────────────────────────────────────────────────
if uploaded_files:
    st.markdown(
        f"<div style='color:#68d391;font-weight:600;margin:0.8rem 0;font-size:0.92rem;'>"
        f"✅ {len(uploaded_files)} PDF file(s) ready for processing</div>",
        unsafe_allow_html=True,
    )

    col_btn, _ = st.columns([3, 5])
    with col_btn:
        run_btn = st.button(
            "🚀 Extract & Classify Transactions",
            use_container_width=True,
        )

    if run_btn or st.session_state.get("pipeline_done"):

        # ── Phase 1: PDF Extraction ──────────────────────────────────────────
        if run_btn:
            st.session_state.pop("pipeline_done", None)
            st.session_state.pop("classified_df", None)

            st.markdown('<div class="step-badge">STEP 2 — Extracting Tables from PDF(s)</div>', unsafe_allow_html=True)

            overall_bar = st.progress(0, text="Reading PDFs…")
            status_ph   = st.empty()

            all_dfs, all_meta, all_page_types = [], {}, {}
            n_files = len(uploaded_files)

            for fi, uploaded_file in enumerate(uploaded_files):
                pdf_bytes = uploaded_file.read()
                file_name = uploaded_file.name
                status_ph.info(f"📄 Parsing **{file_name}** ({fi + 1}/{n_files})…")
                page_ph = st.empty()

                def page_progress(page_idx, total, _fi=fi, _fn=file_name):
                    frac = (_fi + page_idx / total) / (n_files + 1)
                    overall_bar.progress(frac, text=f"{_fn} — Page {page_idx + 1}/{total}")
                    page_ph.markdown(
                        f"<div style='color:#7a9eba;font-size:0.8rem;'>Scanning page {page_idx+1} of {total}…</div>",
                        unsafe_allow_html=True,
                    )

                df_out, meta_out, page_types = extract_bank_statement(
                    pdf_bytes, dpi=dpi_val, progress_cb=page_progress
                )
                if not df_out.empty:
                    df_out.insert(0, "Source File", file_name)
                    all_dfs.append(df_out)
                all_meta.update(meta_out)
                all_page_types[file_name] = page_types
                page_ph.empty()

            if not all_dfs:
                overall_bar.empty()
                status_ph.empty()
                st.warning("⚠️ No transactions found. Check that the PDFs contain structured bank statement tables.")
                st.stop()

            combined_raw = pd.concat(all_dfs, ignore_index=True)

            # ── Phase 2: Auto-Classify Descriptions + Save Embeddings ────────
            overall_bar.progress(0.75, text="🧠 Classifying transactions with Nomic AI…")
            status_ph.info("🏷️ Cleaning descriptions, generating embeddings & classifying 5 domain categories…")

            clf_model, sub_embs = load_classifier_engine()
            source_label = uploaded_files[0].name if n_files == 1 else "combined_statements"

            # Detect description column
            desc_col = None
            for col in combined_raw.columns:
                if str(col).strip().lower() in ["description", "narration", "particulars", "transaction remarks", "details"]:
                    desc_col = col
                    break
            if desc_col is None and not combined_raw.empty:
                desc_col = combined_raw.columns[0]

            classified_df, emb_path, excel_path = classify_and_save_artifacts(
                df=combined_raw,
                source_name=source_label,
                output_base_dir="output",
                model=clf_model,
                sub_embeddings=sub_embs,
                desc_col=desc_col,
            )

            overall_bar.progress(1.0, text="✅ All done!")
            status_ph.empty()

            st.session_state["classified_df"]    = classified_df
            st.session_state["all_meta"]         = all_meta
            st.session_state["saved_emb_path"]   = emb_path
            st.session_state["saved_excel_path"] = excel_path
            st.session_state["pipeline_done"]    = True

        # ── Retrieve from session ────────────────────────────────────────────
        classified_df    = st.session_state.get("classified_df", pd.DataFrame())
        all_meta         = st.session_state.get("all_meta", {})
        saved_emb_path   = st.session_state.get("saved_emb_path", "")
        saved_excel_path = st.session_state.get("saved_excel_path", "")

        if classified_df.empty:
            st.warning("No data available.")
            st.stop()

        # ── STEP 3: Summary Cards ────────────────────────────────────────────
        st.markdown('<div class="step-badge">STEP 3 — Results & AI Classification Summary</div>', unsafe_allow_html=True)

        def parse_amt(s):
            try:
                return float(str(s).replace("Rs.", "").replace(",", "").strip())
            except Exception:
                return 0.0

        total_txns   = len(classified_df)
        debit_count  = int((classified_df.get("Debit",  pd.Series([""])) != "").sum())
        credit_count = int((classified_df.get("Credit", pd.Series([""])) != "").sum())
        total_debit  = sum(parse_amt(v) for v in classified_df.get("Debit",  []) if v)
        total_credit = sum(parse_amt(v) for v in classified_df.get("Credit", []) if v)

        mode_dist  = classified_df["Transaction Mode"].value_counts().to_dict()  if "Transaction Mode"  in classified_df.columns else {}
        purp_dist  = classified_df["Transaction Purpose"].value_counts().to_dict() if "Transaction Purpose" in classified_df.columns else {}

        # Row 1: transaction cards
        c1, c2, c3, c4, c5 = st.columns(5)
        for col, val, label in [
            (c1, str(total_txns),                       "Total Transactions"),
            (c2, str(debit_count),                      "Debit Entries"),
            (c3, str(credit_count),                     "Credit Entries"),
            (c4, f"₹{total_debit:,.0f}",                "Total Debited (₹)"),
            (c5, f"₹{total_credit:,.0f}",               "Total Credited (₹)"),
        ]:
            with col:
                st.markdown(
                    f"<div class='metric-card'>"
                    f"<div class='metric-value'>{val}</div>"
                    f"<div class='metric-label'>{label}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

        st.markdown("<br>", unsafe_allow_html=True)

        # Row 2: classification distribution pills
        if mode_dist or purp_dist:
            d1, d2 = st.columns(2)
            with d1:
                st.markdown(
                    "<div style='background:rgba(26,42,64,0.7);border:1px solid rgba(99,179,237,0.15);"
                    "border-radius:12px;padding:0.9rem 1rem;'>"
                    "<div style='color:#90cdf4;font-weight:700;font-size:0.8rem;letter-spacing:0.06em;margin-bottom:0.6rem;'>"
                    "💳 TRANSACTION MODE BREAKDOWN</div>" +
                    "".join(
                        f"<div style='display:flex;justify-content:space-between;align-items:center;"
                        f"padding:0.3rem 0;border-bottom:1px solid rgba(99,179,237,0.06);'>"
                        f"<span style='color:#c8d8ea;font-size:0.82rem;'>{m}</span>"
                        f"<span style='background:rgba(99,179,237,0.15);color:#90cdf4;font-size:0.78rem;"
                        f"font-weight:700;padding:2px 10px;border-radius:20px;'>{cnt}</span>"
                        f"</div>"
                        for m, cnt in sorted(mode_dist.items(), key=lambda x: -x[1])
                    ) +
                    "</div>",
                    unsafe_allow_html=True,
                )
            with d2:
                st.markdown(
                    "<div style='background:rgba(26,42,64,0.7);border:1px solid rgba(99,179,237,0.15);"
                    "border-radius:12px;padding:0.9rem 1rem;'>"
                    "<div style='color:#90cdf4;font-weight:700;font-size:0.8rem;letter-spacing:0.06em;margin-bottom:0.6rem;'>"
                    "🎯 TRANSACTION PURPOSE BREAKDOWN</div>" +
                    "".join(
                        f"<div style='display:flex;justify-content:space-between;align-items:center;"
                        f"padding:0.3rem 0;border-bottom:1px solid rgba(99,179,237,0.06);'>"
                        f"<span style='color:#c8d8ea;font-size:0.82rem;'>{p}</span>"
                        f"<span style='background:rgba(104,211,145,0.15);color:#68d391;font-size:0.78rem;"
                        f"font-weight:700;padding:2px 10px;border-radius:20px;'>{cnt}</span>"
                        f"</div>"
                        for p, cnt in sorted(purp_dist.items(), key=lambda x: -x[1])
                    ) +
                    "</div>",
                    unsafe_allow_html=True,
                )

        st.markdown("<br>", unsafe_allow_html=True)

        # Artifact saved notice
        if saved_emb_path:
            st.markdown(
                f"<div style='background:rgba(20,46,30,0.7);border:1px solid rgba(104,211,145,0.25);"
                f"border-radius:10px;padding:0.7rem 1rem;margin-bottom:1rem;font-size:0.83rem;'>"
                f"<span style='color:#68d391;font-weight:700;'>📁 Artifacts auto-saved:</span> "
                f"<span style='color:#9ae6b4;'>Embeddings → <code>{saved_emb_path}</code> &nbsp;|&nbsp; "
                f"Excel → <code>{saved_excel_path}</code></span>"
                f"</div>",
                unsafe_allow_html=True,
            )

        # ── Account metadata cards ────────────────────────────────────────────
        if all_meta:
            bank_name = all_meta.get("Bank Name", "Bank Statement")
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:12px;margin:1.5rem 0 0.8rem 0;'>"
                f"<div class='section-header' style='margin:0;'>📋 Account Information</div>"
                f"<span style='background:rgba(99,179,237,0.12);border:1px solid rgba(99,179,237,0.3);"
                f"color:#90cdf4;padding:3px 12px;border-radius:20px;font-size:0.8rem;font-weight:700;'>🏦 {bank_name}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

            ACCOUNT_FIELDS = ["Account Holder", "Account Number", "Customer ID", "Account Type"]
            BRANCH_FIELDS  = ["Branch Name", "IFSC Code", "MICR Code"]
            STMT_FIELDS    = ["Statement Period", "Opening Balance", "Closing Balance"]

            def _frow(lbl, val):
                return (
                    f"<div style='display:flex;justify-content:space-between;padding:0.4rem 0.7rem;"
                    f"border-bottom:1px solid rgba(99,179,237,0.07);'>"
                    f"<span style='color:#7a9eba;font-size:0.8rem;min-width:130px;'>{lbl}</span>"
                    f"<span style='color:#c8d8ea;font-size:0.84rem;font-weight:600;text-align:right;'>{val}</span>"
                    f"</div>"
                )

            def _scard(title, icon, fields):
                rows = "".join(_frow(f, all_meta[f]) for f in fields if all_meta.get(f))
                if not rows:
                    return ""
                return (
                    f"<div style='background:linear-gradient(135deg,#162030,#1a2d44);border:1px solid "
                    f"rgba(99,179,237,0.15);border-radius:12px;overflow:hidden;'>"
                    f"<div style='background:rgba(43,108,176,0.18);padding:0.55rem 0.8rem;font-size:0.78rem;"
                    f"font-weight:700;color:#90cdf4;letter-spacing:0.05em;'>{icon} {title}</div>{rows}</div>"
                )

            mc1, mc2, mc3 = st.columns(3)
            for col, title, icon, fields in [
                (mc1, "ACCOUNT DETAILS", "👤", ACCOUNT_FIELDS),
                (mc2, "BRANCH & CODES",  "🏦", BRANCH_FIELDS),
                (mc3, "STATEMENT",       "📊", STMT_FIELDS),
            ]:
                card = _scard(title, icon, fields)
                if card:
                    with col:
                        st.markdown(card, unsafe_allow_html=True)

            address_val = all_meta.get("Client Address") or all_meta.get("Address", "")
            if address_val:
                st.markdown(
                    f"<div style='background:linear-gradient(135deg,#162030,#1a2d44);border:1px solid "
                    f"rgba(99,179,237,0.15);border-radius:12px;padding:0.75rem 0.9rem;margin-top:0.7rem;'>"
                    f"<div style='font-size:0.76rem;font-weight:700;color:#90cdf4;margin-bottom:0.3rem;'>📍 ADDRESS</div>"
                    f"<div style='color:#c8d8ea;font-size:0.85rem;line-height:1.5;'>{address_val}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            st.markdown("<br>", unsafe_allow_html=True)

        # ── STEP 4: Classified Transaction Ledger ────────────────────────────
        st.markdown('<div class="step-badge">STEP 4 — Classified Transaction Ledger</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-header">All Transactions with 5 AI-Classified Domain Columns</div>', unsafe_allow_html=True)

        display_cols = [c for c in classified_df.columns if c != "Source File"]

        # Search bar
        search_q = st.text_input(
            "🔍 Search transactions:",
            placeholder="Search by description, mode, purpose, counterparty…",
            label_visibility="visible",
        )

        view_df = classified_df[display_cols]
        if search_q.strip():
            mask = view_df.astype(str).apply(
                lambda row: row.str.contains(search_q.strip(), case=False, regex=False).any(), axis=1
            )
            view_df = view_df[mask]

        # 3 tabs only: All / Debits / Credits
        t_all, t_deb, t_cr = st.tabs(["📋 All Transactions", "🔴 Debits Only", "🟢 Credits Only"])

        debit_col  = "Debit"  if "Debit"  in classified_df.columns else None
        credit_col = "Credit" if "Credit" in classified_df.columns else None

        with t_all:
            st.caption(f"Showing {len(view_df)} of {total_txns} transactions")
            st.dataframe(view_df, use_container_width=True, hide_index=True, height=480)

        with t_deb:
            if debit_col:
                deb_df = classified_df[classified_df[debit_col] != ""][display_cols].reset_index(drop=True)
                st.caption(f"{len(deb_df)} debit transactions")
                st.dataframe(deb_df, use_container_width=True, hide_index=True, height=440)
            else:
                st.info("No debit column found.")

        with t_cr:
            if credit_col:
                cr_df = classified_df[classified_df[credit_col] != ""][display_cols].reset_index(drop=True)
                st.caption(f"{len(cr_df)} credit transactions")
                st.dataframe(cr_df, use_container_width=True, hide_index=True, height=440)
            else:
                st.info("No credit column found.")

        # ── STEP 5: Download ─────────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="step-badge">STEP 5 — Download Classified Statement</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-header">Export — All 5 Domain Columns Included (No Scores)</div>', unsafe_allow_html=True)

        excel_bytes = build_styled_excel(classified_df, all_meta)

        dl1, dl2, _ = st.columns([2, 2, 4])
        with dl1:
            st.download_button(
                label="⬇️ Download Excel (.xlsx)",
                data=excel_bytes,
                file_name="classified_bank_statement.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dl2:
            csv_bytes = classified_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="⬇️ Download CSV (.csv)",
                data=csv_bytes,
                file_name="classified_bank_statement.csv",
                mime="text/csv",
                use_container_width=True,
            )

        st.success(
            f"🎉 **{total_txns} transactions** extracted, classified into 5 domain categories, and ready to download!"
        )

# ─────────────────────────────────────────────────────────────────────────────
# Empty State (no file uploaded yet)
# ─────────────────────────────────────────────────────────────────────────────
else:
    st.markdown(
        """
        <div style="text-align:center;padding:4rem 2rem;background:rgba(10,20,35,0.5);
            border:1px dashed rgba(99,179,237,0.25);border-radius:18px;margin-top:1.5rem;">
            <div style="font-size:4rem;margin-bottom:1.2rem;">📂</div>
            <div style="font-size:1.2rem;font-weight:700;color:#63b3ed;margin-bottom:0.5rem;">
                Upload a Bank Statement PDF to Begin
            </div>
            <div style="font-size:0.88rem;color:#4a6e8a;line-height:1.8;max-width:520px;margin:0 auto;">
                The pipeline will automatically:<br>
                ① Extract all transaction rows from the PDF<br>
                ② Verify running balance mathematically<br>
                ③ Clean and embed transaction descriptions (saved to <code>output/embeddings/</code>)<br>
                ④ Classify each transaction into 5 AI domains<br>
                ⑤ Produce a downloadable Excel with all 5 columns appended
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────────────────────────────────────
# Footer
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    "<div class='footer-text'>"
    "BankLens AI · Hybrid Engine: Camelot (digital) + RapidOCR (scanned) · Nomic Embed Text v1.5 · 5-Class Semantic Classifier"
    "</div>",
    unsafe_allow_html=True,
)

